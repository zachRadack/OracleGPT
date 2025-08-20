import os
import re
import json
import glob
from openpyxl import load_workbook
from selenium_extractor import SeleniumExtractor, save_quarterly_patch_data

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')

def get_excel_files():
	"""Discover all Excel files in the data directory"""
	excel_patterns = ['*.xlsx', '*.xls']
	excel_files = []
	
	for pattern in excel_patterns:
		files = glob.glob(os.path.join(DATA_DIR, pattern))
		# Get just the filename, not the full path
		excel_files.extend([os.path.basename(f) for f in files])
	
	return sorted(excel_files)


# Extracts display text and hyperlink from a cell using openpyxl's hyperlink property
def extract_hyperlink(cell):
	if cell is None:
		return None, None
	display_text = cell.value
	url = cell.hyperlink.target if cell.hyperlink else None
	return display_text, url


def process_excel(file_path):
	wb = load_workbook(file_path, data_only=True)
	ws = wb.active
	results = []
	for row in ws.iter_rows(min_row=2):  # Assuming first row is header
		Module_cell = row[0]  # Column C (cell object)
		feature_cell = row[2]  # Column C (cell object)
		delivered_enabled = row[3].value  # Column D
		impact_existing = row[4].value  # Column E
		feature, hyperlink = extract_hyperlink(feature_cell)
		if feature or hyperlink:
			results.append({
				"Module": Module_cell.value if Module_cell.value else None,
				"feature": feature,
				"hyperlink": hyperlink,
				"delivered_enabled": delivered_enabled,
				"impact_to_existing_processes": impact_existing
			})
	return results

def main():
	# Discover all Excel files in the data directory
	excel_files = get_excel_files()
	
	if not excel_files:
		print("No Excel files found in the data directory.")
		return
	
	print(f"Found {len(excel_files)} Excel file(s): {', '.join(excel_files)}")
	
	all_results = []
	for excel_file in excel_files:
		file_path = os.path.join(DATA_DIR, excel_file)
		if os.path.exists(file_path):
			print(f"Processing: {excel_file}")
			try:
				results = process_excel(file_path)
				all_results.extend(results)
				print(f"Extracted {len(results)} features from {excel_file}")
			except Exception as e:
				print(f"Error processing {excel_file}: {e}")
				continue
	
	if not all_results:
		print("No data extracted from Excel files.")
		return
	
	# Write to JSON
	output_path = os.path.join(DATA_DIR, 'features.json')
	with open(output_path, 'w', encoding='utf-8') as f:
		json.dump(all_results, f, indent=2, ensure_ascii=False)
	print(f"Extracted {len(all_results)} total features to {output_path}")
	
	# Now extract content using Selenium
	print("\nStarting Selenium extraction...")
	try:
		with SeleniumExtractor(headless=True) as extractor:
			quarterly_data = extractor.extract_multiple_urls(all_results)
		
		# Save the quarterly patch data
		save_quarterly_patch_data(quarterly_data, DATA_DIR)
		print("Selenium extraction completed successfully!")
		
	except Exception as e:
		print(f"Error during Selenium extraction: {e}")
		print("Features.json has been created, but Selenium extraction failed.")

if __name__ == "__main__":
	main()
